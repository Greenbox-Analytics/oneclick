"""Tests for the expense-report export endpoint.

Covers:
- GET /expenses/export?format=pdf   → PDF StreamingResponse
- GET /expenses/export?format=xlsx  → XLSX StreamingResponse
- invalid format                    → 400
- project_id / category / artist_id filters narrow the rows included in the report

``get_expenses_summary`` is monkeypatched at the router module level so the
tests exercise the export/report path without wiring the full Supabase
membership → projects → expenses chain.
"""

from unittest.mock import AsyncMock

import pytest

PDF_MEDIA = "application/pdf"
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

ROWS = [
    {
        "id": "e1",
        "project_id": "p1",
        "project_name": "Album One",
        "artist_id": "a1",
        "artist_name": "Artist One",
        "description": "Studio time",
        "amount": 100.0,
        "currency": "USD",
        "amount_usd": 100.0,
        "category": "studio",
        "incurred_on": "2026-06-01",
        "is_tagged": False,
    },
    {
        # EUR row: original amount differs from the USD-converted total contribution.
        "id": "e2",
        "project_id": "p2",
        "project_name": "Album Two",
        "artist_id": "a2",
        "artist_name": "Artist Two",
        "description": "Marketing push",
        "amount": 250.0,
        "currency": "EUR",
        "amount_usd": 275.0,
        "category": "marketing",
        "incurred_on": "2026-06-10",
        "is_tagged": True,
    },
    {
        "id": "e3",
        "project_id": "p1",
        "project_name": "Album One",
        "artist_id": "a1",
        "artist_name": "Artist One",
        "description": "Uncategorized misc",
        "amount": 40.0,
        "currency": "USD",
        "amount_usd": 40.0,
        "category": None,
        "incurred_on": None,
        "is_tagged": False,
    },
]


@pytest.fixture()
def _patch_summary(monkeypatch):
    """Patch get_expenses_summary in the router to return ROWS."""
    import expenses.router as exp_router

    monkeypatch.setattr(exp_router, "get_expenses_summary", AsyncMock(return_value=ROWS))


class TestExportFormats:
    def test_pdf_returns_200_and_content_type(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export", params={"format": "pdf"})
        assert resp.status_code == 200
        assert resp.headers["content-type"] == PDF_MEDIA
        assert len(resp.content) > 0

    def test_pdf_is_the_default_format(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == PDF_MEDIA

    def test_xlsx_returns_200_and_content_type(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export", params={"format": "xlsx"})
        assert resp.status_code == 200
        assert resp.headers["content-type"] == XLSX_MEDIA
        assert len(resp.content) > 0

    def test_content_disposition_is_attachment(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export", params={"format": "xlsx"})
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert "Expense_Report_" in cd
        assert ".xlsx" in cd

    def test_invalid_format_returns_400(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export", params={"format": "csv"})
        assert resp.status_code == 400


class TestExportFilters:
    def test_project_filter_scopes_the_report(self, client, mock_supabase, _patch_summary):
        """A project filter narrows both the rows and the scope label in the filename."""
        resp = client.get("/expenses/export", params={"format": "pdf", "project_id": "p1"})
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        # scope label derives from the selected project's name
        assert "Album_One" in cd

    def test_overall_report_uses_all_projects_label(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export", params={"format": "pdf"})
        cd = resp.headers.get("content-disposition", "")
        assert "All_projects" in cd

    def test_category_filter_matches_uncategorized_as_other(self, client, mock_supabase, _patch_summary):
        """category=other includes the null-category row (parity with the page)."""
        resp = client.get("/expenses/export", params={"format": "xlsx", "category": "other"})
        assert resp.status_code == 200
        assert len(resp.content) > 0

    def test_artist_filter_scopes_the_report(self, client, mock_supabase, _patch_summary):
        """An artist filter (without a project) uses the artist's name as the scope label."""
        resp = client.get("/expenses/export", params={"format": "pdf", "artist_id": "a1"})
        assert resp.status_code == 200
        cd = resp.headers.get("content-disposition", "")
        assert "Artist_One" in cd

    def test_project_label_wins_over_artist_label(self, client, mock_supabase, _patch_summary):
        resp = client.get("/expenses/export", params={"format": "pdf", "project_id": "p1", "artist_id": "a1"})
        cd = resp.headers.get("content-disposition", "")
        assert "Album_One" in cd


class TestCurrencyHandling:
    def test_xlsx_has_currency_columns_and_usd_total(self, client, mock_supabase, _patch_summary):
        """Line items keep their original currency; the TOTAL row sums amount_usd."""
        import io

        from openpyxl import load_workbook

        resp = client.get("/expenses/export", params={"format": "xlsx"})
        assert resp.status_code == 200

        ws = load_workbook(io.BytesIO(resp.content))["Expenses"]
        header_row = next(row for row in ws.iter_rows(values_only=True) if row and row[0] == "Date")
        assert "Amount" in header_row
        assert "Currency" in header_row
        assert "Amount (USD)" in header_row

        last_row = list(ws.iter_rows(values_only=True))[-1]
        assert "TOTAL" in last_row
        # 100 USD + 275 USD (from 250 EUR) + 40 USD
        assert last_row[-1] == 415.0

    def test_usd_totals_use_amount_usd(self):
        from expenses.common import grand_total, usd_amount

        assert grand_total(ROWS) == 415.0
        assert usd_amount(ROWS[1]) == 275.0
        # Legacy rows without amount_usd fall back to amount.
        assert usd_amount({"amount": 12.5}) == 12.5

    def test_fmt_money_symbols(self):
        from expenses.common import fmt_money

        assert fmt_money(1234.5, "USD") == "US$1,234.50"
        assert fmt_money(1234.5, "EUR") == "€1,234.50"
        # Unknown currencies fall back to a code suffix.
        assert fmt_money(10, "KES") == "10.00 KES"


class TestFilterExpenseRows:
    """Unit tests for the shared row filter (row scoping isn't visible via the endpoint)."""

    def test_artist_filter_selects_matching_rows(self):
        from expenses.common import filter_expense_rows

        out = filter_expense_rows(ROWS, None, None, "a1")
        assert [r["id"] for r in out] == ["e1", "e3"]

    def test_artist_and_category_intersect(self):
        from expenses.common import filter_expense_rows

        out = filter_expense_rows(ROWS, None, "other", "a1")
        assert [r["id"] for r in out] == ["e3"]
        assert filter_expense_rows(ROWS, None, "marketing", "a1") == []

    def test_rows_without_artist_excluded_when_artist_filter_set(self):
        from expenses.common import filter_expense_rows

        rows = ROWS + [{**ROWS[0], "id": "e4", "artist_id": None, "artist_name": None}]
        assert all(r["artist_id"] == "a1" for r in filter_expense_rows(rows, None, None, "a1"))
        assert [r["id"] for r in filter_expense_rows(rows, None, None, None)] == [
            "e1",
            "e2",
            "e3",
            "e4",
        ]
