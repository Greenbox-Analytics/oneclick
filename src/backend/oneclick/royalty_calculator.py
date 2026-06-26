"""
Royalty Payment Calculator
Calculates payments from royalty statements and music contracts

Installation:
pip install openpyxl openai python-dotenv PyMuPDF

Usage:
    from royalty_calculator import RoyaltyCalculator

    calculator = RoyaltyCalculator()
    payments = calculator.calculate_payments("contract.pdf", "statement.xlsx")
"""

import csv
import difflib
import json
import logging
import os
import re
import time
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from knowledge.reference_search import search_reference
from utils.contract_parsing.basis_detection import audit_contract_basis, log_basis_finding
from utils.contract_parsing.models import ContractData
from utils.contract_parsing.parser import (
    NON_STREAMING_PAYOR_CONTEXT_PHRASES,
    NON_STREAMING_PAYOR_TERMS,
    STREAMING_EQUIVALENT_TERMS,
    MusicContractParser,
)
from utils.llm.client import get_openai_client
from utils.text.normalize import find_matching_song, normalize_name, normalize_title, simplify_role

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()


def is_streaming_equivalent_royalty_type(royalty_type: str) -> bool:
    """Treat master/producer-style royalties as streaming-equivalent payouts."""
    if not royalty_type:
        return False
    normalized = royalty_type.lower()
    # Lowercase each term too — STREAMING_EQUIVALENT_TERMS contains mixed-case
    # entries like "DPD" that would otherwise never match a lowercased input.
    return any(term.lower() in normalized for term in STREAMING_EQUIVALENT_TERMS)


# Build payor-context regexes by substituting each payor name (word-boundary
# anchored, alternation-grouped) into each phrase template. Word boundaries
# prevent short-acronym false positives (e.g. "BMI" inside "BMG"). Phrase
# templates require a payment-direction context so bare mentions like
# "applies to Direct Monies such as SoundExchange" do not trip the denylist.
# Built once at import.
_PAYOR_ALTERNATION = "|".join(re.escape(term) for term in NON_STREAMING_PAYOR_TERMS)
_PAYOR_GROUP = rf"\b(?:{_PAYOR_ALTERNATION})\b"
_NON_STREAMING_PAYOR_PATTERNS = [
    re.compile(phrase.format(payor=_PAYOR_GROUP), re.IGNORECASE) for phrase in NON_STREAMING_PAYOR_CONTEXT_PHRASES
]


def is_streaming_earnable_share(share) -> bool:
    """A share is streaming-earnable iff its royalty_type matches the
    streaming-equivalent allowlist AND its terms do not reference a
    direct-pay collector (SoundExchange, PROs, MLC, etc.) in a payment-
    direction context (e.g. "payable by X", "via X", "X Letter of Direction").

    Bare mentions of a payor name without payment-direction context are NOT
    enough to exclude a share — that avoids false positives on phrasings
    like "applies to Direct Monies (e.g., SoundExchange)" where the payor
    is named as an income-source example, not the payor of the share.
    """
    if not is_streaming_equivalent_royalty_type(share.royalty_type):
        return False
    haystack = f"{share.royalty_type or ''} {share.terms or ''}"
    return not any(p.search(haystack) for p in _NON_STREAMING_PAYOR_PATTERNS)


class CalculationError(ValueError):
    """Raised when a royalty calculation cannot produce results. Carries a
    machine-readable code, a user-facing message, an actionable suggestion,
    and optional structured details for the frontend to render.

    Subclasses ValueError so existing broad catchers and pytest.raises(ValueError)
    assertions continue to work.
    """

    def __init__(self, code: str, message: str, suggestion: str, details: dict | None = None):
        self.code = code
        self.user_message = message
        self.suggestion = suggestion
        self.details = details or {}
        super().__init__(f"[{code}] {message}")


# Statement-songs preview cap for NO_SONG_MATCHES details payload.
_STATEMENT_SONGS_PREVIEW_CAP = 20


@dataclass
class RoyaltyPayment:
    """Represents a calculated royalty payment"""

    song_title: str
    party_name: str
    role: str
    royalty_type: str
    percentage: float
    total_royalty: float
    amount_to_pay: float
    terms: str | None = None


@dataclass
class StatementRow:
    """A single line item from a royalty statement, preserving the dimensional
    fields (vendor, country, delivery format, sale date) that the song-level
    aggregator discards. Powers the OneClick Earnings Breakdown tab.

    `sale_date` is normalized to an ISO `YYYY-MM-DD` string, or None if the
    statement's date couldn't be parsed.
    """

    song_title: str
    net_payable: float
    vendor: str | None = None
    country: str | None = None
    country_code: str | None = None
    delivery_type: str | None = None
    delivery_format: str | None = None
    sale_date: str | None = None
    units_sold: float | None = None
    net_units: float | None = None
    sales: float | None = None
    net_income: float | None = None
    distribution: float | None = None
    isrc: str | None = None
    upc: str | None = None
    currency: str = "USD"


# Normalized statement header -> StatementRow field. Headers are lowercased and
# stripped before lookup; the first matching alias for a field wins.
_STATEMENT_DIMENSION_ALIASES: dict[str, list[str]] = {
    "vendor": ["vendor"],
    "country": ["country of sale", "country"],
    "country_code": ["country code"],
    "delivery_type": ["delivery type"],
    "delivery_format": ["delivery format"],
    "sale_date": ["sale date"],
    "units_sold": ["units sold"],
    "net_units": ["net units"],
    "sales": ["sales"],
    "net_income": ["net income"],
    "distribution": ["distribution"],
    "isrc": ["isrc"],
    "upc": ["upc"],
    "currency": ["currency"],
}

# Numeric StatementRow fields parsed best-effort (commas stripped; junk -> None).
_STATEMENT_NUMERIC_FIELDS = {
    "units_sold",
    "net_units",
    "sales",
    "net_income",
    "distribution",
}


def _coerce_iso_date(raw) -> str | None:
    """Normalize a statement date cell to ISO `YYYY-MM-DD`, or None if unparseable.

    Accepts `M/D/YYYY` (distributor reports) and `YYYY-MM-DD`. A datetime/date
    value (openpyxl may hand one back) is formatted directly.
    """
    if raw is None or raw == "":
        return None
    # openpyxl with data_only may return a datetime/date directly.
    if hasattr(raw, "strftime"):
        try:
            return raw.strftime("%Y-%m-%d")
        except (ValueError, AttributeError):
            return None
    s = str(raw).strip()
    if not s:
        return None
    from datetime import datetime

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _coerce_number(raw) -> float | None:
    """Best-effort float coercion; strips thousands separators. None on failure."""
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


class RoyaltyCalculator:
    """
    Enhanced calculator for royalty payments from statements and contracts.

    Features:
    - Auto-detects columns in royalty statements
    - Fuzzy matching between contract works and statement songs
    - Support for single or multiple contracts
    - Intelligent merging of multiple contracts
    - Comprehensive validation and reporting
    """

    def __init__(self, api_key: str | None = None):
        """
        Initialize the calculator with OpenAI API key.

        Args:
            api_key: Optional OpenAI API key. If not provided, will look in environment variables.
        """
        # Load API key from parameter or environment
        self.api_key = api_key or os.getenv("OPENAI_API_KEY")

        if not self.api_key or not self.api_key.startswith("sk-"):
            raise ValueError("❌ OpenAI API key missing or invalid. Please add it to .env file or pass as parameter.")

        # Initialize contract parser with explicit API key
        self.contract_parser = MusicContractParser(api_key=self.api_key)

    # ========================================================================
    # ROYALTY STATEMENT READING
    # ========================================================================

    def read_royalty_statement(
        self, excel_path: str, title_column: str | None = None, payable_column: str | None = None
    ) -> dict[str, float]:
        """
        Read streaming royalty statement and calculate total per song.

        Supports both CSV and Excel files. Auto-detects column names if not specified.

        Args:
            excel_path: Path to the royalty statement (CSV or Excel)
            title_column: Name of column containing song titles (auto-detects if None)
            payable_column: Name of column containing net payable amounts (auto-detects if None)

        Returns:
            Dictionary mapping song titles to total net payable amounts
        """
        try:
            logger.info(f"\n📊 Reading royalty statement: {Path(excel_path).name}")

            # Detect file type
            file_ext = Path(excel_path).suffix.lower()

            if file_ext == ".csv":
                return self._read_csv_statement(excel_path, title_column, payable_column)
            elif file_ext in [".xlsx", ".xls"]:
                return self._read_excel_statement(excel_path, title_column, payable_column)
            else:
                raise CalculationError(
                    code="STATEMENT_UNSUPPORTED_FORMAT",
                    message=f"We can't read {file_ext} files for royalty statements.",
                    suggestion="Upload the statement as a CSV or Excel (.xlsx / .xls) file.",
                )

        except CalculationError:
            # Pass structured errors through unchanged so the endpoint can route
            # them to the user with their reason code intact.
            raise
        except Exception as e:
            raise Exception(f"Error reading royalty statement: {str(e)}")

    def read_royalty_statement_rows(self, statement_path: str) -> list[StatementRow]:
        """Read a royalty statement preserving per-line-item dimensions.

        Unlike `read_royalty_statement` (which sums to song-level totals), this
        returns one `StatementRow` per statement line, carrying vendor, country,
        delivery format, and sale date — the fields the Earnings Breakdown tab
        aggregates over. Rows missing a title or net-payable amount are skipped;
        a bad sale date downgrades to None rather than dropping the row.

        Supports CSV and Excel. Title and payable columns are auto-detected with
        the same helpers the aggregator uses.
        """
        file_ext = Path(statement_path).suffix.lower()
        if file_ext == ".csv":
            normalized_rows = self._iter_csv_rows(statement_path)
        elif file_ext in [".xlsx", ".xls"]:
            normalized_rows = self._iter_excel_rows(statement_path)
        else:
            raise CalculationError(
                code="STATEMENT_UNSUPPORTED_FORMAT",
                message=f"We can't read {file_ext} files for royalty statements.",
                suggestion="Upload the statement as a CSV or Excel (.xlsx / .xls) file.",
            )

        normalized_rows = list(normalized_rows)
        if not normalized_rows:
            return []

        headers = list(normalized_rows[0].keys())
        title_key = self._find_title_column(headers)
        payable_key = self._find_payable_column(headers)

        # Resolve each StatementRow field to a concrete header present in the file.
        field_to_key: dict[str, str] = {}
        for field, aliases in _STATEMENT_DIMENSION_ALIASES.items():
            for alias in aliases:
                if alias in headers:
                    field_to_key[field] = alias
                    break

        rows: list[StatementRow] = []
        for raw in normalized_rows:
            title_cell = raw.get(title_key)
            title = str(title_cell).strip() if title_cell not in (None, "") else ""
            net_payable = _coerce_number(raw.get(payable_key))
            if not title or net_payable is None:
                continue

            values: dict = {}
            for field, key in field_to_key.items():
                cell = raw.get(key)
                if field == "sale_date":
                    values[field] = _coerce_iso_date(cell)
                elif field in _STATEMENT_NUMERIC_FIELDS:
                    values[field] = _coerce_number(cell)
                elif field == "currency":
                    values[field] = str(cell).strip() if cell not in (None, "") else "USD"
                else:
                    values[field] = str(cell).strip() if cell not in (None, "") else None

            rows.append(StatementRow(song_title=title, net_payable=net_payable, **values))

        return rows

    def _iter_csv_rows(self, csv_path: str) -> list[dict]:
        """Read a CSV into a list of {normalized_header: value} dicts."""
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                return []
            field_map = {h: h.strip().lower() for h in reader.fieldnames}
            return [{field_map[h]: row.get(h) for h in reader.fieldnames} for row in reader]

    def _iter_excel_rows(self, excel_path: str) -> list[dict]:
        """Read an Excel sheet into a list of {normalized_header: value} dicts."""
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            workbook.close()
            return []
        headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
        result = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            result.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
        workbook.close()
        return result

    def _read_csv_statement(
        self, csv_path: str, title_column: str | None = None, payable_column: str | None = None
    ) -> dict[str, float]:
        """Read royalty statement from CSV file"""
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)

            # Get headers (keys from DictReader)
            headers = [h.strip().lower() for h in reader.fieldnames]

            logger.info(f"   Found {len(headers)} columns: {', '.join(headers[:5])}...")

            # Auto-detect columns if not specified
            if title_column is None:
                title_column = self._find_title_column(headers)
                logger.info(f"   ✓ Auto-detected title column: '{title_column}'")
            else:
                title_column = title_column.lower()

            if payable_column is None:
                payable_column = self._find_payable_column(headers)
                logger.info(f"   ✓ Auto-detected payable column: '{payable_column}'")
            else:
                payable_column = payable_column.lower()

            # Find original column names (case-sensitive)
            original_headers = {h.strip().lower(): h for h in reader.fieldnames}
            title_col_original = original_headers.get(title_column)
            payable_col_original = original_headers.get(payable_column)

            if not title_col_original or not payable_col_original:
                raise CalculationError(
                    code="STATEMENT_COLUMNS_UNDETECTABLE",
                    message="We couldn't find the song-title or net-payable columns in your statement.",
                    suggestion=(
                        "Rename your columns to something we recognize — e.g. 'Title' / 'Song' for "
                        "the title column, and 'Net Payable' / 'Net Royalty' for the amount."
                    ),
                    details={
                        "looking_for": [title_column, payable_column],
                        "available_columns": list(original_headers.values()),
                    },
                )

            # Read data and sum by song title
            song_totals = {}
            rows_processed = 0
            rows_skipped = 0

            for row in reader:
                title_val = row.get(title_col_original, "").strip()
                payable_val = row.get(payable_col_original, "")

                if title_val and payable_val:
                    try:
                        # Handle comma-separated numbers (e.g., "1,555")
                        payable_clean = payable_val.replace(",", "")
                        amount = float(payable_clean)
                        song_totals[title_val] = song_totals.get(title_val, 0.0) + amount
                        rows_processed += 1
                    except (ValueError, TypeError):
                        rows_skipped += 1
                        continue
                else:
                    rows_skipped += 1

            logger.info(f"   ✓ Processed {rows_processed} rows ({rows_skipped} skipped)")
            logger.info(f"   ✓ Found {len(song_totals)} unique songs")

            if not song_totals:
                raise CalculationError(
                    code="STATEMENT_EMPTY",
                    message="We couldn't read any earnings from the royalty statement.",
                    suggestion=(
                        "Check that you uploaded the right file. The statement should be a CSV "
                        "or Excel with one row per song earnings entry."
                    ),
                )

            return song_totals

    def _read_excel_statement(
        self, excel_path: str, title_column: str | None = None, payable_column: str | None = None
    ) -> dict[str, float]:
        """Read royalty statement from Excel file"""
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = workbook.active

        # Extract headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())

        logger.info(f"   Found {len(headers)} columns: {', '.join(headers[:5])}...")

        # Auto-detect columns if not specified
        if title_column is None:
            title_column = self._find_title_column(headers)
            logger.info(f"   ✓ Auto-detected title column: '{title_column}'")
        else:
            title_column = title_column.lower()

        if payable_column is None:
            payable_column = self._find_payable_column(headers)
            logger.info(f"   ✓ Auto-detected payable column: '{payable_column}'")
        else:
            payable_column = payable_column.lower()

        # Find column indices
        try:
            title_idx = headers.index(title_column)
            payable_idx = headers.index(payable_column)
        except ValueError:
            raise CalculationError(
                code="STATEMENT_COLUMNS_UNDETECTABLE",
                message="We couldn't find the song-title or net-payable columns in your statement.",
                suggestion=(
                    "Rename your columns to something we recognize — e.g. 'Title' / 'Song' for "
                    "the title column, and 'Net Payable' / 'Net Royalty' for the amount."
                ),
                details={
                    "looking_for": [title_column, payable_column],
                    "available_columns": list(headers),
                },
            )

        # Read data and sum by song title
        song_totals = {}
        rows_processed = 0
        rows_skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[title_idx] and row[payable_idx] is not None:
                title = str(row[title_idx]).strip()
                try:
                    amount = float(row[payable_idx])
                    song_totals[title] = song_totals.get(title, 0.0) + amount
                    rows_processed += 1
                except (ValueError, TypeError):
                    rows_skipped += 1
                    continue
            else:
                rows_skipped += 1

        workbook.close()

        logger.info(f"   ✓ Processed {rows_processed} rows ({rows_skipped} skipped)")
        logger.info(f"   ✓ Found {len(song_totals)} unique songs")

        if not song_totals:
            raise CalculationError(
                code="STATEMENT_EMPTY",
                message="We couldn't read any earnings from the royalty statement.",
                suggestion=(
                    "Check that you uploaded the right file. The statement should be a CSV "
                    "or Excel with one row per song earnings entry."
                ),
            )

        return song_totals

    def _find_title_column(self, headers: list[str]) -> str:
        """Auto-detect the title column from headers"""
        title_variations = [
            "release title",
            "title",
            "song title",
            "track title",
            "song name",
            "track name",
            "release name",
            "track",
            "song",
        ]

        for header in headers:
            header_clean = header.lower().strip()
            for var in title_variations:
                if var == header_clean or var in header_clean:
                    return header

        raise CalculationError(
            code="STATEMENT_COLUMNS_UNDETECTABLE",
            message="We couldn't identify which column in your statement holds the song titles.",
            suggestion=(
                "Rename your title column to something we recognize — e.g. 'Title', 'Song', "
                "'Track Title', or 'Release Title'."
            ),
            details={"available_columns": list(headers)},
        )

    def _find_payable_column(self, headers: list[str]) -> str:
        """
        Auto-detect the net payable column from headers using 3 layers:
        1. Keyword matching (exact/partial)
        2. Fuzzy matching
        3. Semantic search (LLM)
        """

        # --- Layer 1: Keyword Matching ---

        # Priority variations (more specific matches first)
        priority_variations = [
            "net payable",
            "net payment",
            "net earnings",
            "net paytotal payable",
            "net revenue",
            "net amount",
            "payable to artist",
            "artist payable",
        ]

        # Check priority variations first
        for header in headers:
            header_clean = header.lower().strip()
            for var in priority_variations:
                if var == header_clean or var in header_clean:
                    # Exclude withheld/deduction columns
                    if "withheld" not in header_clean and "deduction" not in header_clean:
                        return header

        # Fallback to general variations
        general_variations = ["payable", "amount", "earnings", "payment", "revenue"]

        for header in headers:
            header_clean = header.lower().strip()
            for var in general_variations:
                if var in header_clean:
                    # Exclude unwanted columns
                    excluded_terms = ["withheld", "deduction", "fee", "commission", "advance"]
                    if not any(term in header_clean for term in excluded_terms):
                        return header

        logger.info("   ⚠️  Layer 1 (Keyword) failed to find payable column. Trying Layer 2 (Fuzzy)...")

        # --- Layer 2: Fuzzy Matching ---

        target_terms = ["net payable", "net amount", "payable", "total payable", "royalty amount", "net pay"]

        # Get close matches for each target term against all headers
        # We use a cutoff of 0.8 for high confidence
        best_match = None
        highest_ratio = 0.0

        for header in headers:
            header_clean = header.lower().strip()
            # Skip likely irrelevant columns to avoid false positives
            if any(
                x in header_clean for x in ["date", "isrc", "upc", "territory", "country", "label", "artist", "title"]
            ):
                continue

            for target in target_terms:
                ratio = difflib.SequenceMatcher(None, header_clean, target).ratio()
                if ratio > highest_ratio and ratio > 0.8:  # 80% similarity threshold
                    highest_ratio = ratio
                    best_match = header

        if best_match:
            logger.info(
                f"   ✓ Layer 2 (Fuzzy) detected payable column: '{best_match}' (confidence: {highest_ratio:.2f})"
            )
            return best_match

        logger.info("   ⚠️  Layer 2 (Fuzzy) failed. Trying Layer 3 (Semantic)...")

        # --- Layer 3: Semantic Search (LLM) ---

        if self.api_key:
            try:
                client = get_openai_client()

                prompt = (
                    f"Given these column headers from a music royalty statement: {headers}\\n\\n"
                    "Identify the single column that represents the 'Net Payable Amount' or 'Royalty Amount' "
                    "that should be paid to the licensor/artist. "
                    "Ignore columns representing gross revenue, fees, taxes, or deductions unless they are the only option.\\n"
                    "Return ONLY the exact column name from the list. If none match, return 'None'."
                )

                response = client.chat.completions.create(
                    model="gpt-4o-mini",  # Use a fast/cheap model
                    messages=[
                        {
                            "role": "system",
                            "content": "You are a data analyst helper. Output only the requested column name.",
                        },
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0.0,
                )

                result = response.choices[0].message.content.strip()

                # Verify the result is actually in the headers
                # (LLM might strip quotes or change case slightly, so we check loosely)
                if result and result.lower() != "none":
                    # Try to find exact match first
                    if result in headers:
                        logger.info(f"   ✓ Layer 3 (Semantic) detected payable column: '{result}'")
                        return result

                    # Try case-insensitive match
                    for h in headers:
                        if h.lower() == result.lower():
                            logger.info(f"   ✓ Layer 3 (Semantic) detected payable column: '{h}'")
                            return h

            except Exception as e:
                logger.warning(f"   ⚠️  Layer 3 (Semantic) error: {e}")

        raise CalculationError(
            code="STATEMENT_COLUMNS_UNDETECTABLE",
            message="We couldn't identify which column in your statement holds the net-payable amounts.",
            suggestion=(
                "Rename your amount column to something we recognize — e.g. 'Net Payable', "
                "'Net Royalty', 'Royalty Payable', or 'Net Income'."
            ),
            details={"available_columns": list(headers)},
        )

    # ========================================================================
    # CONTRACT MERGING
    # ========================================================================

    def merge_contracts(self, contracts: list[ContractData]) -> ContractData:
        """
        Intelligently merge multiple ContractData objects.

        Handles:
        - Name normalization and deduplication
        - Role preservation
        - Royalty share conflict resolution
        - Summary combination

        Args:
            contracts: List of ContractData objects to merge

        Returns:
            Single merged ContractData object
        """
        if not contracts:
            raise ValueError("No contracts provided to merge")

        if len(contracts) == 1:
            return contracts[0]

        logger.info(f"\n🔄 Merging {len(contracts)} contracts...")

        merged_parties = []
        merged_works = []
        merged_royalty_shares = []
        summaries = []

        # Track seen items for deduplication
        seen_parties = {}  # normalized_name -> Party
        seen_works = {}  # normalized_title -> Work
        _seen_shares = {}  # (party, type) -> RoyaltyShare

        for idx, contract in enumerate(contracts, 1):
            logger.info(f"   Processing contract {idx}/{len(contracts)}...")

            if contract.contract_summary:
                summaries.append(contract.contract_summary)

            # Merge parties
            for party in contract.parties:
                norm_name = normalize_name(party.name)
                if norm_name:
                    if norm_name not in seen_parties:
                        seen_parties[norm_name] = party
                        merged_parties.append(party)
                    else:
                        existing = seen_parties[norm_name]
                        # Combine roles from both contracts
                        existing_roles = set(r.strip() for r in existing.role.split(";"))
                        new_roles = set(r.strip() for r in party.role.split(";"))
                        combined = existing_roles | new_roles
                        # Remove generic 'party' if a specific role exists
                        if len(combined) > 1:
                            combined.discard("party")
                        existing.role = "; ".join(sorted(combined))

            # Merge works
            for work in contract.works:
                norm_title = normalize_title(work.title)
                if norm_title:
                    if norm_title not in seen_works:
                        seen_works[norm_title] = work
                        merged_works.append(work)
                    else:
                        # Update work_type if current one is more specific
                        existing = seen_works[norm_title]
                        if work.work_type != "work" and existing.work_type == "work":
                            existing.work_type = work.work_type

            # Merge royalty shares with conflict resolution
            for share in contract.royalty_shares:
                norm_name = normalize_name(share.party_name)

                # Check for existing share with same name AND similar type
                existing_share_for_party = None

                # Determine if current share is streaming-related
                is_streaming_share = is_streaming_equivalent_royalty_type(share.royalty_type)

                for existing in merged_royalty_shares:
                    if normalize_name(existing.party_name) == norm_name:
                        # Check if percentages match
                        if abs(existing.percentage - share.percentage) < 0.01:
                            # CRITICAL: Only consider it a duplicate if the royalty TYPE is also similar.
                            # If one is "Publishing" and one is "Streaming", they are different entitlements
                            # even if they have the same percentage (e.g. 50% Pub / 50% Master).

                            is_existing_streaming = is_streaming_equivalent_royalty_type(existing.royalty_type)

                            # If both are streaming or both are NOT streaming (e.g. both publishing), likely a duplicate
                            if is_streaming_share == is_existing_streaming:
                                existing_share_for_party = existing
                                break

                if existing_share_for_party:
                    logger.info(
                        f"      ℹ️  Duplicate share found for {share.party_name} ({share.percentage}%) - skipping"
                    )
                    continue

                # If no exact duplicate found, add it
                merged_royalty_shares.append(share)

        # Simplify combined roles
        for party in merged_parties:
            party.role = simplify_role(party.role)

        # Combine summaries
        merged_summary = "\n\n".join([s for s in summaries if s.strip()])

        logger.info(
            f"   ✓ Merged to: {len(merged_parties)} parties, {len(merged_works)} works, {len(merged_royalty_shares)} royalty entries"
        )

        return ContractData(
            parties=merged_parties,
            works=merged_works,
            royalty_shares=merged_royalty_shares,
            contract_summary=merged_summary if merged_summary else None,
        )

    # ========================================================================
    # PAYMENT CALCULATION
    # ========================================================================

    def calculate_payments(
        self,
        contract_path: str,
        statement_path: str,
        user_id: str = None,
        contract_id: str = None,
        title_column: str | None = None,
        payable_column: str | None = None,
        full_text: str = None,
    ) -> list[RoyaltyPayment]:
        """
        Calculate payments for single contract and statement.

        Args:
            contract_path: Path to the contract file (not used, kept for compatibility)
            statement_path: Path to the royalty statement (Excel)
            user_id: User ID for querying Pinecone
            contract_id: Contract ID for querying Pinecone
            title_column: Optional - Name of title column in statement
            payable_column: Optional - Name of payable column in statement

        Returns:
            List of RoyaltyPayment objects with calculated amounts
        """
        logger.info("\n" + "=" * 80)
        logger.info("ROYALTY PAYMENT CALCULATION")
        logger.info("=" * 80)

        total_start = time.time()

        # Step 1: Parse contract using full document text
        logger.info("\n📄 Step 1: Extracting contract data from full document...")
        t0 = time.time()
        contract_data = self.contract_parser.parse_contract(full_text=full_text)
        logger.info(f"⏱️  Step 1 took: {time.time() - t0:.2f}s")

        # Step 2: Read royalty statement
        logger.info("\n💵 Step 2: Reading royalty statement...")
        t0 = time.time()
        song_totals = self.read_royalty_statement(statement_path, title_column, payable_column)
        logger.info(f"⏱️  Step 2 took: {time.time() - t0:.2f}s")

        # Step 3: Calculate payments
        logger.info("\n🔍 DEBUG CHECK — Contract vs Statement")
        logger.info("Contract works:")
        for w in contract_data.works:
            logger.info(f"   → {w.title}")
        logger.info("\nStatement songs:")
        for s in list(song_totals.keys())[:10]:
            logger.info(f"   → {s}")

        t0 = time.time()
        payments = self._calculate_payments_from_data(contract_data, song_totals)
        logger.info(f"⏱️  Step 3 took: {time.time() - t0:.2f}s")

        # Step 4 (log-only): detect an explicit basis/deduction clause and record it for review.
        # Does NOT change any payout. Never breaks the calc.
        if full_text:
            try:
                types_present = [s.royalty_type for s in contract_data.royalty_shares]
                finding = audit_contract_basis(
                    full_text,
                    types_present,
                    openai_client=self.contract_parser.openai_client,
                    search_fn=search_reference,
                )
                payments = log_basis_finding(payments, finding, contract_id=contract_id or "unknown", user_id=user_id)
            except Exception as e:  # noqa: BLE001 — nuance logging must never break the payout calc
                logger.warning(f"[NuanceAudit] skipped (detection/logging error): {e}")

        logger.info(f"\n✅ Total calculation process took: {time.time() - total_start:.2f}s")

        return payments

    def calculate_payments_from_contract_ids(
        self,
        contract_ids: list[str],
        user_id: str,
        statement_path: str,
        title_column: str | None = None,
        payable_column: str | None = None,
        contract_markdowns: dict[str, str] = None,
    ) -> list[RoyaltyPayment]:
        """
        Parse multiple contracts from Pinecone in PARALLEL, merge their data, and calculate payments.

        Args:
            contract_ids: List of contract IDs to query from Pinecone
            user_id: User ID for Pinecone namespace
            statement_path: Path to the royalty statement file
            title_column: Optional column name for song titles
            payable_column: Optional column name for payable amounts

        Returns:
            List of RoyaltyPayment objects with combined results
        """
        logger.info("\n" + "=" * 80)
        logger.info(f"MULTI-CONTRACT ROYALTY CALCULATION ({len(contract_ids)} contracts)")
        logger.info("=" * 80)

        # Step 1: Parse all contracts in PARALLEL
        logger.info(f"\n📄 Step 1: Parsing {len(contract_ids)} contracts (Parallel)...")
        all_contracts_data = []

        from concurrent.futures import ThreadPoolExecutor, as_completed

        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_cid = {
                executor.submit(
                    self.contract_parser.parse_contract,
                    full_text=contract_markdowns.get(cid) if contract_markdowns else None,
                ): cid
                for cid in contract_ids
            }

            # Process results as they complete
            for _i, future in enumerate(as_completed(future_to_cid), 1):
                cid = future_to_cid[future]
                try:
                    # logger.info(f"   ...Finished processing a contract...")
                    data = future.result()
                    all_contracts_data.append(data)
                    logger.info(
                        f"   ✓ Contract parsed successfully ({len(data.parties)} parties, {len(data.works)} works)"
                    )
                except Exception as e:
                    logger.error(f"   ⚠️  Failed to parse contract {cid}: {e}")

        if not all_contracts_data:
            raise ValueError("❌ No valid contracts could be parsed. Please check your files.")

        # Step 2: Merge contracts
        merged_data = self.merge_contracts(all_contracts_data)

        # Step 3: Read royalty statement
        logger.info("\n💵 Step 2: Reading royalty statement...")
        song_totals = self.read_royalty_statement(statement_path, title_column, payable_column)

        # Step 4: Calculate payments
        payments = self._calculate_payments_from_data(merged_data, song_totals)
        logger.info("[NuanceAudit] multi-contract: basis-nuance detection deferred (v1 single-contract only)")

        return payments

    def calculate_payments_from_contracts(
        self,
        contract_paths: list[str],
        statement_path: str,
        title_column: str | None = None,
        payable_column: str | None = None,
    ) -> list[RoyaltyPayment]:
        """
        Parse multiple contracts, merge their data, and calculate payments.

        This is useful when multiple contracts cover the same works with
        different contributors (e.g., producer agreement, artist agreement).

        Args:
            contract_paths: List of paths to contract files
            statement_path: Path to the royalty statement file
            title_column: Optional column name for song titles
            payable_column: Optional column name for payable amounts

        Returns:
            List of RoyaltyPayment objects with combined results
        """
        logger.info("\n" + "=" * 80)
        logger.info(f"MULTI-CONTRACT ROYALTY CALCULATION ({len(contract_paths)} contracts)")
        logger.info("=" * 80)

        # Step 1: Parse all contracts
        logger.info(f"\n📄 Step 1: Parsing {len(contract_paths)} contracts...")
        all_contracts_data = []

        for idx, path in enumerate(contract_paths, 1):
            try:
                logger.info(f"\n   Contract {idx}/{len(contract_paths)}: {Path(path).name}")
                data = self.contract_parser.parse_contract(path)
                all_contracts_data.append(data)

                # Quick preview
                logger.info(
                    f"      → {len(data.parties)} parties, {len(data.works)} works, {len(data.royalty_shares)} shares"
                )

            except Exception as e:
                logger.error(f"      ⚠️  Failed to parse: {e}")
                continue

        if not all_contracts_data:
            raise ValueError("❌ No valid contracts could be parsed. Please check your files.")

        # Step 2: Merge contracts
        merged_data = self.merge_contracts(all_contracts_data)

        # Step 3: Read royalty statement
        logger.info("\n💵 Step 2: Reading royalty statement...")
        song_totals = self.read_royalty_statement(statement_path, title_column, payable_column)

        # Step 4: Calculate payments
        payments = self._calculate_payments_from_data(merged_data, song_totals)

        return payments

    def _calculate_payments_from_data(
        self, contract_data: ContractData, song_totals: dict[str, float]
    ) -> list[RoyaltyPayment]:
        """
        Internal method to calculate payments from parsed contract data.

        Args:
            contract_data: Parsed ContractData object
            song_totals: Dictionary of song titles to amounts

        Returns:
            List of RoyaltyPayment objects
        """
        logger.info("\n💰 Step 3: Calculating payments...\n")

        # Validate inputs
        if not contract_data.works:
            raise CalculationError(
                code="NO_WORKS_IN_CONTRACT",
                message="We couldn't find any songs in the contract you uploaded.",
                suggestion=(
                    "Double-check the contract file. Make sure it's the right document and "
                    "lists the songs / compositions / masters it covers."
                ),
            )

        if not contract_data.royalty_shares:
            raise CalculationError(
                code="NO_ROYALTY_SHARES_IN_CONTRACT",
                message="We couldn't find any royalty splits in the contract.",
                suggestion=(
                    "Re-check the contract — look for a Schedule, Exhibit, or table listing "
                    "percentage splits. If splits aren't expressed as percentages, royalties "
                    "can't be calculated automatically."
                ),
            )

        if not song_totals:
            raise CalculationError(
                code="STATEMENT_EMPTY",
                message="We couldn't read any earnings from the royalty statement.",
                suggestion=(
                    "Check that you uploaded the right file. The statement should be a CSV "
                    "or Excel with one row per song earnings entry."
                ),
            )

        # Debug: log raw royalty_type vs terms to disambiguate streaming-filter misses
        logger.info("   🔎 Royalty shares before streaming filter:")
        for i, share in enumerate(contract_data.royalty_shares, 1):
            logger.info(f"      {i}. royalty_type={share.royalty_type!r}")
            logger.info(f"         terms={share.terms!r}")

        # Filter for streaming-earnable shares: must pass the streaming-equivalent
        # allowlist AND not reference a direct-pay collector (SoundExchange, PROs,
        # MLC) whose royalties are paid outside DSP streaming statements.
        streaming_shares = [share for share in contract_data.royalty_shares if is_streaming_earnable_share(share)]

        excluded_payor_count = sum(
            1
            for share in contract_data.royalty_shares
            if is_streaming_equivalent_royalty_type(share.royalty_type) and not is_streaming_earnable_share(share)
        )
        if excluded_payor_count:
            logger.info(
                f"   ↪️  Excluded {excluded_payor_count} share(s) tied to direct-pay "
                f"collectors (SoundExchange / PRO / MLC) — paid outside DSP statements"
            )

        if not streaming_shares:
            logger.warning("   ⚠️  No streaming royalty shares found in contract")
            raise CalculationError(
                code="NO_STREAMING_EARNABLE_SHARES",
                message="The contract has royalty shares, but none of them are earned through streaming.",
                suggestion=(
                    "OneClick calculates streaming royalties. If the contract covers publishing, "
                    "mechanical, sync, or SoundExchange royalties, those are paid through different "
                    "statements — try uploading the matching statement type, or use a contract that "
                    "covers streaming/master royalties."
                ),
                details={"excluded_payor_count": excluded_payor_count},
            )

        logger.info(f"   Found {len(streaming_shares)} streaming royalty shares")
        logger.info(f"   Found {len(contract_data.works)} works to match")

        # Calculate payments for each work
        payments = []
        matched_count = 0
        unmatched_works = []

        for work in contract_data.works:
            # Find matching song in statement
            matching_song, total_royalty = find_matching_song(work.title, song_totals)

            if matching_song:
                matched_count += 1
                logger.info(f"\n   ✓ '{work.title}'")
                logger.info(f"      Matched to: '{matching_song}'")
                logger.info(f"      Total royalties: ${total_royalty:,.2f}")

                # Calculate payment for each party with streaming shares
                for share in streaming_shares:
                    amount_to_pay = total_royalty * (share.percentage / 100.0)

                    # Find party details
                    party = next(
                        (
                            p
                            for p in contract_data.parties
                            if normalize_name(p.name) == normalize_name(share.party_name)
                        ),
                        None,
                    )
                    role = party.role if party else "unknown"

                    payment = RoyaltyPayment(
                        song_title=work.title,
                        party_name=share.party_name,
                        role=role,
                        royalty_type=share.royalty_type,
                        percentage=share.percentage,
                        total_royalty=total_royalty,
                        amount_to_pay=amount_to_pay,
                        terms=share.terms,
                    )
                    payments.append(payment)

                    logger.info(f"         → {share.party_name} ({role}): {share.percentage}% = ${amount_to_pay:,.2f}")
            else:
                unmatched_works.append(work.title)

        # Summary
        logger.info("\n   📊 Matching Summary:")
        logger.info(f"      ✓ Matched: {matched_count}/{len(contract_data.works)} works")

        if unmatched_works:
            logger.warning("      ⚠️  Unmatched works:")
            for title in unmatched_works:
                logger.warning(f"         - {title}")
            logger.info("\n      💡 Tip: Check for typos or verify these songs are in the statement")

        if matched_count == 0:
            statement_song_titles = list(song_totals.keys())
            raise CalculationError(
                code="NO_SONG_MATCHES",
                message="The contract covers songs that don't appear in this royalty statement.",
                suggestion=(
                    "Make sure the statement is for the same release as the contract. If a song "
                    "title is spelled differently in each, edit one to match. The list below shows "
                    "what we found in each."
                ),
                details={
                    "contract_works": [work.title for work in contract_data.works],
                    "statement_songs": statement_song_titles[:_STATEMENT_SONGS_PREVIEW_CAP],
                    "statement_song_total_count": len(statement_song_titles),
                },
            )

        logger.info(f"\n   ✅ Calculated {len(payments)} total payments")

        return payments

    # ========================================================================
    # OUTPUT METHODS
    # ========================================================================

    def save_payments_to_excel(self, payments: list[RoyaltyPayment], output_path: str):
        """
        Save calculated payments to an Excel file with formatting.

        Args:
            payments: List of RoyaltyPayment objects
            output_path: Path to save the Excel file
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Royalty Payments"

        # Headers
        headers = [
            "Song Title",
            "Payee Name",
            "Role",
            "Royalty Type",
            "Share %",
            "Total Song(s) Revenue",
            "Amount Owed",
            "Terms",
        ]
        sheet.append(headers)

        # Format headers
        _header_font = openpyxl.styles.Font(bold=True, size=11)
        header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF")

        for cell in sheet[1]:
            cell.font = header_font_white
            cell.fill = header_fill

        # Add data
        for payment in payments:
            sheet.append(
                [
                    payment.song_title,
                    payment.party_name,
                    payment.role,
                    payment.royalty_type,
                    f"{payment.percentage}%",
                    payment.total_royalty,
                    payment.amount_to_pay,
                    payment.terms or "",
                ]
            )

        # Format currency columns
        for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7):
            for cell in row:
                cell.number_format = "$#,##0.00"

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 3, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Add summary row
        if payments:
            summary_row = sheet.max_row + 2
            sheet[f"A{summary_row}"] = "TOTAL"
            sheet[f"A{summary_row}"].font = openpyxl.styles.Font(bold=True)

            total_formula = f"=SUM(G2:G{sheet.max_row - 1})"
            sheet[f"G{summary_row}"] = total_formula
            sheet[f"G{summary_row}"].font = openpyxl.styles.Font(bold=True)
            sheet[f"G{summary_row}"].number_format = "$#,##0.00"

        workbook.save(output_path)
        logger.info(f"\n💾 Payment breakdown saved to {output_path}")

    def save_payments_to_json(self, payments: list[RoyaltyPayment], output_path: str):
        """
        Save calculated payments to a JSON file.

        Args:
            payments: List of RoyaltyPayment objects
            output_path: Path to save the JSON file
        """
        data = {
            "payments": [asdict(payment) for payment in payments],
            "summary": {
                "total_payments": len(payments),
                "total_amount": sum(p.amount_to_pay for p in payments),
                "unique_payees": len(set(p.party_name for p in payments)),
                "unique_songs": len(set(p.song_title for p in payments)),
            },
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info(f"💾 Payment data saved to {output_path}")

    def print_payment_summary(self, payments: list[RoyaltyPayment]):
        """Print a formatted summary of payments to console"""

        logger.info("\n" + "=" * 80)
        logger.info("PAYMENT SUMMARY")
        logger.info("=" * 80)

        if not payments:
            logger.warning("\n⚠️  No payments calculated")
            return

        # Group by payee
        payee_totals = {}
        for payment in payments:
            if payment.party_name not in payee_totals:
                payee_totals[payment.party_name] = {"role": payment.role, "total": 0.0, "details": []}
            payee_totals[payment.party_name]["total"] += payment.amount_to_pay
            payee_totals[payment.party_name]["details"].append(payment)

        # Print summary for each payee
        for payee, data in sorted(payee_totals.items()):
            logger.info(f"\n👤 {payee} ({data['role'].title()})")
            logger.info(f"   Total Payment: ${data['total']:,.2f}")
            logger.info("   Breakdown:")

            for detail in data["details"]:
                logger.info(
                    f"      • {detail.song_title}: "
                    f"{detail.percentage}% of ${detail.total_royalty:,.2f} "
                    f"= ${detail.amount_to_pay:,.2f}"
                )

        # Grand total
        grand_total = sum(p.amount_to_pay for p in payments)
        logger.info(f"\n{'=' * 80}")
        logger.info(f"GRAND TOTAL: ${grand_total:,.2f}")
        logger.info(f"Total Payments: {len(payments)}")
        logger.info(f"Unique Payees: {len(payee_totals)}")
        logger.info(f"Unique Songs: {len(set(p.song_title for p in payments))}")
        logger.info(f"{'=' * 80}\n")
