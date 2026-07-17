"""Expense-report Excel generator (openpyxl).

Writes to an ``io.BytesIO`` buffer (not a file path) so the router can stream it
directly. Sheet 1 is the itemized list with a TOTAL row; sheet 2 is the
by-category summary.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from expenses.common import category_label, category_totals, grand_total, sorted_rows

HEADER_FILL = PatternFill(start_color="1A3A2A", end_color="1A3A2A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
MONEY_FMT = "#,##0.00"


def generate_expense_report_xlsx(
    rows: list[dict],
    scope_label: str,
    category_label_str: str,
    generated_on: str,
) -> io.BytesIO:
    wb = Workbook()

    # --- Expenses sheet ---
    ws = wb.active
    ws.title = "Expenses"

    ws.append([f"Expense Report — {scope_label}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([f"Category: {category_label_str}"])
    ws.append([f"Generated: {generated_on}  ·  Amounts in USD"])
    ws.append([])

    headers = ["Date", "Project", "Artist", "Description", "Category", "Amount (USD)"]
    ws.append(headers)
    header_idx = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    amount_col = len(headers)
    for r in sorted_rows(rows):
        ws.append(
            [
                r.get("incurred_on") or "",
                r.get("project_name") or "",
                r.get("artist_name") or "",
                r.get("description") or "",
                category_label(r.get("category")),
                round(float(r.get("amount") or 0), 2),
            ]
        )

    total = round(grand_total(rows), 2)
    ws.append(["", "", "", "", "TOTAL", total])
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = BOLD

    # Amount column: number format + right alignment for data + total rows
    for row in range(header_idx + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=amount_col)
        cell.number_format = MONEY_FMT
        cell.alignment = Alignment(horizontal="right")

    for i, width in enumerate([12, 24, 20, 40, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Summary sheet (by category) ---
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Category", "Total (USD)"])
    for col in (1, 2):
        cell = ws2.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for cat, amount in category_totals(rows):
        ws2.append([category_label(cat), round(amount, 2)])
    ws2.append(["TOTAL", total])
    ws2.cell(row=ws2.max_row, column=1).font = BOLD
    ws2.cell(row=ws2.max_row, column=2).font = BOLD
    for row in range(2, ws2.max_row + 1):
        ws2.cell(row=row, column=2).number_format = MONEY_FMT
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
